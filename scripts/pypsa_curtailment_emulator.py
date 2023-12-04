# -*- coding: utf-8 -*-
"""
Created on 29th of November 2023
Author: Ebbe Kyhl Gøtske

This script contains the function "create_emulator" which collects
the results from the parameterization steps and updates the parameters to 
the PyPSA-Eur emulator excel file. 
Furthermore, it contains the functions used for testing and visualization of 
the emulator estimations at a user-specified scenario (currently, defined by 
the wind and solar penetration level as well as the rollout of short-duration 
and long-duration electricity storage).
"""

from openpyxl import load_workbook
from pycel import ExcelCompiler
import matplotlib.pyplot as plt
from scripts.plotting import read_and_plot_scenarios
from scripts.curtailment_parameterization import base_curtailment, technology_term, convert_series_into_2D_matrix
import numpy as np
import pandas as pd

def create_emulator(scenarios, RDIR, file_path,efficiencies,tech_labels, update_excel_file=False):
    variables = ["wind curt", # For parameterization step 1&2
                "solar curt",  # For parameterization step 1&2
                "wind absolute curt", # For parameterization step 1&2
                "solar absolute curt",  # For parameterization step 1&2
                "SDES dispatch", # For parameterization step 2
                "SDES absolute dispatch", # For parameterization step 2
                "SDES energy cap", # For parameterization step 2
                "SDES discharge cap", # For parameterization step 2
                "LDES dispatch", # For parameterization step 2
                "LDES absolute dispatch", # For parameterization step 2
                "LDES energy cap", # For parameterization step 2
                "LDES discharge cap", # For parameterization step 2
                "system cost", # For parameterization step 4 
                "cv wind", # For paramterization step 3
                "cv solar", # For paramterization step 3
                "backup capacity", # For paramterization step 3&4
                "backup capacity factor high", # For paramterization step 3 
                "backup capacity factor low", # For paramterization step 3
                "total demand",
                "transmission volume"
                ]

    base_case = scenarios["base"]
    ################## BASE CURTAILMENT ##############################

    df = read_and_plot_scenarios([base_case],
                                    variables_input = variables, 
                                    directory=RDIR,
                                    plot=False)

    demand = df[base_case]["total demand"].iloc[0]*1e6 # total demand in MWh - here assuming that the demand is the same for all scenarios!
    bin_lvls = [0,0.1,0.3,0.4,0.5,0.6,0.7,0.9]

    # Wind curtailment
    var = 'wind absolute curt' # variable 
    x_name = "wind" # primary index
    gamma_ij_wind_series, x_share_df = base_curtailment(df,var,bin_lvls,demand,x_name,base_case) # base curtailment parameters
    print("Wind curtailment successfully parameterized! Proceeding to solar curtailment...")

    # save the parameters to the MESSAGEix-GLOBIOM subfolder (for later use):
    x_share_df.to_csv("MESSAGEix_GLOBIOM/wind_shares.csv") # share of electricity
    gamma_ij_wind_series.to_csv("MESSAGEix_GLOBIOM/gamma_ij_wind.csv") # save series to .csv

    # convert gamma_ij_wind_series into 2D array format:
    gamma_ij_wind = convert_series_into_2D_matrix(gamma_ij_wind_series,
                                                    lvls=[0,1,2,3,4,5,6],
                                                    x_i_str="wind_curtailment_w",
                                                    x_j_str="solar")

    # Solar curtailment
    var = 'solar absolute curt' # variable
    x_name = "solar" # primary index
    gamma_ij_solar_series, x_share_df = base_curtailment(df,var,bin_lvls,demand,x_name,base_case) # base curtailment parameters
    print("Solar curtailment successfully parameterized! Proceeding to technology term...")

    # save to .csv:
    x_share_df.to_csv("MESSAGEix_GLOBIOM/solar_shares.csv") # share of electricity
    gamma_ij_solar_series.to_csv("MESSAGEix_GLOBIOM/gamma_ij_solar.csv") # save series to .csv

    # convert the series into a matrix for visualisation purposes:
    gamma_ij_solar = convert_series_into_2D_matrix(gamma_ij_solar_series,
                                                    lvls=[0,1,2,3,4,5,6],
                                                    x_i_str="solar_curtailment_s",
                                                    x_j_str="wind")

    ########################## Technology term ##############################

    renewables = ["wind","solar"]
    techs = efficiencies.keys()

    for renewable in renewables:
        for tech in techs:
            scen_base = scenarios[tech][renewable][0] # base case
            scen_ref = scenarios[tech][renewable][1] # reference case without technology 
            scen_tech = scenarios[tech][renewable][2] # reference case with technology

            scens = [scen_base,
                    scen_ref,
                    scen_tech,
                    ]
            
            df_tech = read_and_plot_scenarios(scens,
                                            variables_input = variables, 
                                            directory=RDIR,
                                            plot=False)

            beta = technology_term(df_tech, 
                                    scen_base, scen_ref, scen_tech, 
                                    tech_name = tech, tech_label = tech_labels[tech], tech_efficiency = efficiencies[tech], 
                                    renewable=renewable)
            
            beta.to_csv("MESSAGEix_GLOBIOM/beta_" + tech + "_" + renewable + ".csv")

            if renewable == "wind" and tech == "LDES":
                beta_wind_ldes = beta.copy()        
                beta_wind_ldes.index = gamma_ij_wind.index
                beta_wind_ldes.columns = gamma_ij_wind.columns

            elif renewable == "solar" and tech == "LDES":
                beta_solar_ldes = beta.copy()
                beta_solar_ldes.index = gamma_ij_solar.index
                beta_solar_ldes.columns = gamma_ij_solar.columns

            elif renewable == "wind" and tech == "SDES":
                beta_wind_sdes = beta.copy()
                beta_wind_sdes.index = gamma_ij_wind.index
                beta_wind_sdes.columns = gamma_ij_wind.columns
            
            elif renewable == "solar" and tech == "SDES":
                beta_solar_sdes = beta.copy()
                beta_solar_sdes.index = gamma_ij_solar.index
                beta_solar_sdes.columns = gamma_ij_solar.columns

            print(tech + " impact on " + renewable + " curtailment successfully parameterized!")

    print("Parameterization done!")

    ############################# Update the excel file #########################################

    if update_excel_file:
        # Load the workbook using openpyxl
        workbook = load_workbook(file_path)

        # Select the desired sheets
        sheet_name_wind = "wind curtailment"
        sheet_name_solar = "solar curtailment"
        sheet_wind = workbook[sheet_name_wind] # sheet with wind curtailment
        sheet_solar = workbook[sheet_name_solar] # sheet with solar curtailment

        # Update the cells with the new parameter values
        for i in range(7):
            for j in range(7):
                # allocate base curtailment
                sheet_wind["D22:J28"][i][j].value = gamma_ij_wind.loc[i][j]
                sheet_solar["D22:J28"][i][j].value = gamma_ij_solar.loc[i][j]

                # allocate technology term for LDES
                sheet_wind["D51:J57"][i][j].value = beta_wind_ldes.loc[i][j]
                sheet_solar["D42:J48"][i][j].value = beta_solar_ldes.loc[i][j]

                # allocate technology term for SDES
                sheet_wind["D42:J48"][i][j].value = beta_wind_sdes.loc[i][j]
                sheet_solar["D51:J57"][i][j].value = beta_solar_sdes.loc[i][j]

        # Save the changes to the workbook
        workbook.save(file_path)

def update_excel_parameters(file_path, sheet_names, inputs):

    # Load the workbook using openpyxl
    workbook = load_workbook(file_path)

    input_values_wind = {'C13': inputs["wind"], 
                         'D13': inputs["solar"],
                         'D15': inputs["sdes"],
                         'D16': inputs["ldes"],
                         }
    
    input_values_solar = {'D13': inputs["wind"], 
                          'C13': inputs["solar"],
                          'D15': inputs["ldes"],
                          'D16': inputs["sdes"],
                          }

    # Select the desired sheets
    sheet_name_wind = sheet_names["wind"]
    sheet_name_solar = sheet_names["solar"]
    sheet_wind = workbook[sheet_name_wind] # sheet with wind curtailment
    sheet_solar = workbook[sheet_name_solar] # sheet with solar curtailment

    # Update the cell with the new parameter value
    # sheet[cell_address] = new_value
    for cell_address, new_value in input_values_wind.items():
        sheet_wind[cell_address] = new_value
        
    for cell_address, new_value in input_values_solar.items():
        sheet_solar[cell_address] = new_value

    # Save the changes to the workbook
    workbook.save(file_path)

def read_excel_outputs(file_path, output_address):

    # Load the workbook using pycel
    workbook = ExcelCompiler(file_path)

    output_addresses = ["wind curtailment" + output_address,
                        "solar curtailment" + output_address]
    # Read output values
    output = {cell_address: workbook.evaluate(cell_address) for cell_address in output_addresses}
    
    output1 = output["wind curtailment" + output_address]
    output2 = output["solar curtailment" + output_address]

    return output1, output2

def run_pypsa_emulator(file_path,curtailment_type,wind_res,solar_res,ldes,sdes):
    
    import time

    sheet_name_wind = "wind curtailment"
    sheet_name_solar = "solar curtailment"

    inputs = {'wind': wind_res,
              'solar': solar_res,
              'ldes': ldes,
              'sdes': sdes,
              }

    sheet_names = {'wind': sheet_name_wind,
                   'solar': sheet_name_solar}

    outputs_addresses = {"demand": '!L37',
                         "resources": '!L38',}
    
    output_address = outputs_addresses[curtailment_type]

    # Update input parameters
    update_excel_parameters(file_path, sheet_names, inputs)

    time.sleep(2) # currently, I need to have this much wait time to avoid data race issues
    
    # Read and print the output value
    result = read_excel_outputs(file_path, output_address)

    return result

def color_plot_2D(series, renewable="", vmax=30, cbar_label="resources", disp=0.5):
    # convert multiindex dataframe to 2D array
    df = series.unstack()
    df = df.iloc[::-1]

    df = df.mask(df == '') # replace empty cells with NaN
    df = df.astype(float) # convert to float
    
    # plot
    cmap = "Reds"
    fig, ax = plt.subplots()
    ax.pcolormesh(df.columns,df.index,df, cmap = cmap, vmin = 0, vmax = vmax)
    ax.set_xlabel("wind share (% of el. demand)")
    ax.set_ylabel("solar PV share (% of el. demand)")

    # add colorbar
    cbar = plt.colorbar(ax.pcolormesh(df.columns,
                                      df.index,df, 
                                      cmap = cmap, 
                                      vmin = 0, 
                                      vmax = vmax))
    cbar.ax.set_ylabel(renewable + ' curtailment (% of ' + cbar_label + ")")

    # add cell values to the plot if the dataframe is not too big (otherwise, it gets messy):
    if len(df.index) <= 15:
        for i in range(len(df.index)):
            for j in range(len(df.columns)):
                c = df.iloc[i,j]
                if c != '' and c != np.nan and c > 0:
                    ax.text(df.columns[j]+disp, df.index[i]+disp, "{:.1f}".format(c), va='center', ha='center')

    return fig, ax, df

def estimate_calculation_time(N):
    # reference timing
    N_ref = 324. # number of iterations if range = (10,100) and stepsize = 5
    timing_ref = 18. # minutes (including 2sec wait time between each iteration)
    calculation_rate = timing_ref/N_ref # seconds per iteration

    timing = calculation_rate*N # timing in minutes

    timing = round(timing,1)
    return timing

def test_pypsa_emulator(file_path, wind_res,solar_res,ldes,sdes, curtailment_type="resources"):
    # wind_res [list] = wind resources in percentage of demand (including curtailment)
    # solar_res [list] = solar PV resources percentage of demand (including curtailment)
    # ldes [scalar] = long-duration energy storage disptach as percentage of demand (including energy losses)
    # sdes [scalar]= short-duration energy storage dispatch as percentage of demand (including energy losses)
    # curtailment_type [string] = curtailment as percentage of demand ("demand") or renewable resources ("resources")

    N = len(wind_res)*len(solar_res)
    timing = estimate_calculation_time(N)
    print("estimated timing: ", timing, " minute")

    wind_curtailment_dict = {}
    solar_curtailment_dict = {}
    for solar_res_i in solar_res:
        for wind_res_i in wind_res:
            op1,op2 = run_pypsa_emulator(file_path,curtailment_type,wind_res_i,solar_res_i,ldes,sdes)

            wind_curtailment_dict[(solar_res_i,wind_res_i)] = op1 # output1 is wind curtailment
            solar_curtailment_dict[(solar_res_i,wind_res_i)] = op2 # output2 is solar curtailment

    wind_curtailment_series = pd.Series(wind_curtailment_dict)
    solar_curtailment_series = pd.Series(solar_curtailment_dict)

    return wind_curtailment_series, solar_curtailment_series